import datetime

from Function.SQLConnect import SQLConnect
from openpyxl import load_workbook


class StudentsManager:
    def __init__(self):
        self.lists = None
        self.sql = SQLConnect()

    def checkID(self, id_number):
        if len(id_number) != 18:
            return False

            # 检查身份证号码的前 17 位是否都是数字
        if not id_number[:-1].isdigit():
            return False

            # 检查身份证号码的最后一位
        check_sum = 0
        factors = [7, 9, 10, 5, 8, 4, 2, 1, 6, 3, 7, 9, 10, 5, 8, 4, 2]
        check_codes = ['1', '0', 'X', '9', '8', '7', '6', '5', '4', '3', '2']
        for i in range(17):
            check_sum += int(id_number[i]) * factors[i]
        if id_number[-1] != check_codes[check_sum % 11]:
            return False

        # 检查出生日期是否有效
        birth_date = id_number[6:14]
        try:
            year = int(birth_date[:4])
            month = int(birth_date[4:6])
            day = int(birth_date[6:])
            if month < 1 or month > 12:
                return False
            if day < 1 or day > 31:
                return False
            if month in [4, 6, 9, 11] and day == 31:
                return False
            if month == 2:
                if year % 400 == 0 or (year % 100 != 0 and year % 4 == 0):
                    if day > 29:
                        return False
                else:
                    if day > 28:
                        return False
        except ValueError:
            return False

        # 如果所有检查都通过了，那么身份证号码的格式是正确的
        return True

    def getSex(self, id_number):
        if int(id_number[16]) % 2 == 0:
            return 1
        else:
            return 0

    def getBirthDate(self, id_number):
        return id_number[6:14]

    def get_list1(self, list1):
        self.lists = []
        for i in list1:
            self.lists.append(i[0])

    def get_list2(self, list2):
        for i in list2:
            self.lists.append(i[1])

    def submit_data(self, lists):
        if self.sql.commit_students_info(lists):
            return True
        else:
            return False

    def load_data(self, filename):
        wb = load_workbook(filename)
        ws = wb.active
        list1 = []
        for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
            for cell in row:
                if cell.value is None:
                    return True
                list1.append(cell.value)
                print(list1)
                if not list1:
                    continue
            if self.submit_data(list1):
                list1.clear()
            else:
                return False
        return True

    def get_students_info(self):
        db = self.sql.connect()
        cursor = db.cursor()
        cursor.execute("SELECT * FROM StudentsBasicInfo")
        rows = cursor.fetchall()
        db.close()
        return rows